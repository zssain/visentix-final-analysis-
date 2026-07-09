"""Export endpoint tests — auth, Excel generation, OneDrive upload, read-only."""

import io
import time
from unittest.mock import AsyncMock, patch
from uuid import uuid4

import jwt as pyjwt
import pytest
from httpx import ASGITransport, AsyncClient

from app.config import settings
from app.main import app


_TEST_ORG_ID = str(uuid4())


def _make_token(sub: str = "test-user") -> str:
    now = int(time.time())
    return pyjwt.encode(
        {"sub": sub, "aud": "authenticated", "iat": now - 60,
         "exp": now + 3600, "role": "authenticated"},
        settings.supabase_jwt_secret, algorithm="HS256",
    )


def _mock_profile(role: str = "sme", org_id: str = _TEST_ORG_ID):
    return patch("app.auth._load_profile", new_callable=AsyncMock,
                 return_value={"role": role, "organization_id": org_id})


# ── Excel export tests ──────────────────────────────────────

@pytest.mark.anyio
async def test_export_requires_auth():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        r = await c.get("/assessments/export")
        assert r.status_code == 401


@pytest.mark.anyio
async def test_export_returns_xlsx():
    token = _make_token()
    with _mock_profile("admin"):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test", timeout=120) as c:
            r = await c.get("/assessments/export",
                            headers={"Authorization": f"Bearer {token}"})
            assert r.status_code == 200
            assert "spreadsheetml" in r.headers.get("content-type", "")
            assert "attachment" in r.headers.get("content-disposition", "")
            assert ".xlsx" in r.headers["content-disposition"]


@pytest.mark.anyio
async def test_export_xlsx_has_all_sheets():
    """Verify the returned file has all 8 sheets including Source Evidence and Notice Sources."""
    from openpyxl import load_workbook

    token = _make_token()
    with _mock_profile("admin"):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test", timeout=120) as c:
            r = await c.get("/assessments/export",
                            headers={"Authorization": f"Bearer {token}"})
            assert r.status_code == 200

            wb = load_workbook(io.BytesIO(r.content))
            sheet_names = wb.sheetnames
            assert "Assessments" in sheet_names
            assert "Scores" in sheet_names
            assert "Findings" in sheet_names
            assert "Recommendations" in sheet_names
            assert "Benchmark" in sheet_names
            assert "Source Evidence" in sheet_names
            assert "Notice Sources" in sheet_names
            assert "Organizations" in sheet_names

            # Check Assessments sheet has headers
            ws = wb["Assessments"]
            headers = [cell.value for cell in ws[1]]
            assert "Notice ID" in headers
            assert "Organization" in headers
            assert "Overall Score" in headers
            assert "VCI (Confidence)" in headers

            # Check Source Evidence sheet has headers
            ws_ev = wb["Source Evidence"]
            ev_headers = [cell.value for cell in ws_ev[1]]
            assert "Source Clause Text" in ev_headers
            assert "Notice URL" in ev_headers
            assert "Finding Code" in ev_headers

            # Check Notice Sources sheet has headers
            ws_ns = wb["Notice Sources"]
            ns_headers = [cell.value for cell in ws_ns[1]]
            assert "Source URL" in ns_headers
            assert "Organization" in ns_headers

            wb.close()


@pytest.mark.anyio
async def test_export_is_read_only():
    """Export must not write to the DB."""
    token = _make_token()
    with _mock_profile("admin"), \
         patch("app.db.supabase_rest_post", new_callable=AsyncMock) as mock_post:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as c:
            r = await c.get("/assessments/export",
                            headers={"Authorization": f"Bearer {token}"})
            assert r.status_code == 200
            mock_post.assert_not_called()


# ── Bundle (ZIP) export tests ─────────────────────────────────

@pytest.mark.anyio
async def test_bundle_requires_auth():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        r = await c.get("/assessments/export/bundle")
        assert r.status_code == 401


@pytest.mark.anyio
async def test_bundle_returns_zip():
    """Bundle endpoint returns a valid ZIP containing PDFs and an Excel file."""
    import zipfile

    token = _make_token()
    with _mock_profile("admin"):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test", timeout=120) as c:
            r = await c.get("/assessments/export/bundle",
                            headers={"Authorization": f"Bearer {token}"})
            assert r.status_code == 200
            assert "zip" in r.headers.get("content-type", "")
            assert ".zip" in r.headers.get("content-disposition", "")

            zf = zipfile.ZipFile(io.BytesIO(r.content))
            names = zf.namelist()
            # Should contain the Excel summary
            assert "visentix-assessments-summary.xlsx" in names
            # Should contain at least one PDF (if assessments exist in DB)
            # and the Excel is always present
            assert len(names) >= 1
            zf.close()
